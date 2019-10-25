import sys
import time
import snap7
from Fun import get_real
import struct
import re
from getpass import getpass

import openpyxl
from openpyxl import workbook

'''
Берется файл Export.xlsx который генерируется Tag Management -> Edit -> Export из WinCC 7.3
Файл распарсивается и из него берется все что нужно.

#TODO   Плюс ко всему этому нужно читать тег по длине указанной в фале. 4 байта, 2 байта, 1 байт. Булевый тег.
#TODO   Обработка сделана только для тегов DD. Нужно сделать и для остальных типов тегов.
Команды использованые в данной консольной программе.
list - просто вывод списка тегов. после команды можно ввести часть названия тега. 
set - операция присваивания. Требуется что бы пароль был введен. set pr1_cd{ENTER} вывод значения тега и потом ввод нужного значения.
pass - для операций присвоения тега потребуется ввод пароля
# - с решетки начинается описание части тега которого нужно просто вывести в консоли
'''

try:
    wb = openpyxl.load_workbook(filename = 'Export.xlsx', read_only=True)
except:
    print('Проверте файл!!!')
    input('Нажмите ENTER.')
    sys.exit()


DB = {}
Connect_Controller = False
Access = False
count_tag = 0
ws1 = wb['Connections']
ws2 = wb['Groups']
ws3 = wb['Tags']

Connections = ws1['a1':'d'+str(ws1.max_row)]

Tag = ws3['a1':'g'+str(ws3.max_row)]

#for i in range(3,ws1.max_row):
#    print(Connections[i][0].value, ' ==> ', Connections[i][3].value.split(',')[1])

while True:
    cmd1 = input('Введите команду >> ')
    command = cmd1.lower().split(' ')

    if command[0] == 'exit' or command[0] == '':
        Connect_Controller = False
        #client.disconnect()
        sys.exit()

    if command[0] == 'pass':
        password = getpass('Введите пароль >> ')
        if password == 'Iddqd1':
            Access = True
            print('Доступ разрешен')

    if len(command) == 2:
        if command[0] == 'list':
            for i in range(3,ws3.max_row):
                if Tag[i][0].value.lower().find(command[1]) >= 0:
                    print(Tag[i][0].value)
        elif command[0] == 'set':
            if Access == True:
                #print('Доступ разрешен')
                for i in range(3,ws3.max_row):
                    if Tag[i][0].value.lower().find(command[1]) >= 0:
                        print(Tag[i][0].value)
                        count_tag += 1
                        find_tag = Tag[i][0].value
                if count_tag == 1:
                    count_tag = 0
                    print('Здесь записываем данные в тег.')
                if count_tag > 1:
                    count_tag = 0
                    print('Выберите нужный тег из списка.')
            else:
                print('Доступ запрещен!')
                #sys.exit()
        else:
            print('Неверная команда!')
    
    #первые символы в адресе могут быть:
    #   DB  I   M   MW   MB     MD
    if command[0][:1] == '#':
        for i in range(3,ws3.max_row):
            #print(command[0][1:])
            if Tag[i][0].value.lower().find(command[0][1:]) >= 0:
                for j in range(3,ws1.max_row):
                    if Tag[i][4].value == Connections[j][0].value:              # Если имя соединения правильное
                        if Tag[i][6].value[:2] == 'DB':                         # Если в названии присутствует ДБ то значит мы сможем его прочитать из ДБ.
                            tDB, tDD = re.findall('(\d+)', Tag[i][6].value)     # пишем во временные переменные номер ДБ и номер ДД
                            if not Connect_Controller:
                                client = snap7.client.Client()
                                stat = client.connect(Connections[j][3].value.split(',')[1], int(Connections[j][3].value.split(',')[3]), int(Connections[j][3].value.split(',')[4]))
                                Connect_Controller = True
                            if Connect_Controller:
                                DB[int(tDB)]     = client.db_read(int(tDB), 0, int(tDD)+4)
                                Data        = get_real(DB[int(tDB)], int(tDD))
                                print(Tag[i][0].value, '\t\t', Data)
        client.disconnect()
        Connect_Controller = False




'''
    if Tag[i][6].value[:2] == 'DB':
        print(re.findall('(\d+)', Tag[i][6].value))          #выход за длинну списка out of range
'''



'''


#z = (1024).to_bytes(4,byteorder='big') # правильная запись integer
#value = 17.77
#ba = bytearray(struct.pack(">f", value)) #правильная запись Float point
#fff = client.db_write(110,976,ba)  #(b'\x00\x00\x00\x00')



'''
